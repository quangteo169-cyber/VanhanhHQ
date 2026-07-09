# -*- coding: utf-8 -*-
"""
Dựng trang "Tổng quan vận hành" (Google Sheets / Excel) từ sheet "Data theo SPDV"
của file Dash_Report_PCU_2026.

Kiến trúc 3 lớp:
  1. 'Data theo SPDV'   — snapshot dữ liệu nguồn (GIỮ NGUYÊN tên sheet + địa chỉ ô,
                          để khi copy 2 sheet dashboard sang file Google Sheet thật,
                          công thức tự bám vào sheet nguồn thật).
  2. 'DB_Tổng quan'     — lớp dữ liệu trung gian cho biểu đồ (đơn vị: triệu ₫),
                          đã áp bộ lọc Từ tháng / Đến tháng.
  3. 'Tổng quan vận hành' — trang hiển thị: bộ lọc + 5 thẻ KPI + 5 biểu đồ + bảng so sánh tháng.

Chạy:  python3 tools/build_tong_quan_dashboard.py <input.xlsx> <output.xlsx>
"""
import sys

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC_SHEET = "Data theo SPDV"
DB_SHEET = "DB_Tổng quan"
DASH_SHEET = "Tổng quan vận hành"

# ---- Vị trí dữ liệu trong 'Data theo SPDV' (hàng Excel 1-based, cột F..Q = tháng 1..12)
ROW_NHAP_TONG = 4          # Tổng CO Nhập
ROWS_NHAP = range(5, 14)   # 9 SPDV
ROW_XUAT_TONG = 17         # Tổng CO Xuất
ROWS_XUAT = range(18, 27)
ROW_TH_TONG = 35           # Thiệt hại
ROWS_TH = range(36, 41)    # 5 nhóm
ROW_TON_TONG = 47          # Tổng CO TỒN
ROWS_TON = range(48, 51)   # 3 nhóm
COL_M1 = 6                 # cột F = tháng 1

# ---- Bảng màu categorical (đã validate CVD bằng dataviz/validate_palette.js — PASS)
# Màu đi theo THỰC THỂ (SPDV), cố định trên mọi biểu đồ.
BLUE, AQUA, YELLOW, GREEN = "2A78D6", "1BAF7A", "EDA100", "008300"
VIOLET, RED, MAGENTA, ORANGE = "4A3AA7", "E34948", "E87BA4", "EB6834"
GRAY = "898781"  # nhóm trung tính / bằng 0 cả năm (Token)

SPDV_COLORS = [BLUE, GRAY, AQUA, YELLOW, GREEN, VIOLET, RED, MAGENTA, ORANGE]
# thứ tự: Robux, Token, Gift Card, Nick Random Roblox, Razer Gold,
#         RBX Daily, RBX HQS, Topup, Supercell
TH_COLORS = [BLUE, YELLOW, GREEN, VIOLET, AQUA]   # Robux die, Nick RR Die, Razer Gold, RBX (phí), GC lỗi
TON_COLORS = [GRAY, AQUA, YELLOW]                 # Item, Gift Card (Hub), Nick Random Roblox

INK = "0B0B0B"
INK2 = "52514E"
CARD_FILL = "F0EFEC"
INPUT_FILL = "E7F0FB"
BORDER_C = "C3C2B7"

thin = Side(style="thin", color=BORDER_C)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def col_m(m):
    """Chữ cột của tháng m (1..12) trong sheet nguồn."""
    return get_column_letter(COL_M1 + m - 1)


def copy_snapshot(src_ws_values, ws):
    """Chép giá trị A1:R50 của 'Data theo SPDV' (snapshot, không công thức)."""
    for r, row in enumerate(src_ws_values, start=1):
        if r > 50:
            break
        for c, v in enumerate(row[:18], start=1):
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                v = "'" + v
            ws.cell(row=r, column=c, value=v)
    ws["A2"] = "SNAPSHOT dữ liệu (giá trị tĩnh) — trong file thật, sheet này là sheet nguồn có sẵn."
    ws["A2"].font = Font(size=9, italic=True, color=INK2)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["D"].width = 30
    for c in range(5, 18):
        ws.column_dimensions[get_column_letter(c)].width = 14
    for r in (ROW_NHAP_TONG, ROW_XUAT_TONG, ROW_TH_TONG, ROW_TON_TONG):
        ws.cell(row=r, column=1).font = Font(bold=True)
    for r in range(3, 51):
        for c in range(5, 18):
            cell = ws.cell(row=r, column=c)
            if r == 3:
                cell.number_format = "mm/yyyy"
            else:
                cell.number_format = "#,##0"


# =====================================================================
# Lớp 2: DB_Tổng quan — các khối dữ liệu biểu đồ (triệu ₫, đã áp bộ lọc)
# =====================================================================
def month_block(ws, title, head_row, names, src_rows, unit_fmt):
    """Khối 12 hàng tháng × N cột SPDV + cột 'Tổng' + hàng 'Tổng kỳ lọc'.

    head_row: hàng tiêu đề cột; dữ liệu ở head_row+1 .. head_row+12; tổng ở head_row+13.
    Trả về (hàng đầu dữ liệu, hàng cuối dữ liệu, hàng tổng).
    """
    ws.cell(row=head_row - 1, column=1, value=title).font = Font(bold=True, size=11, color=INK)
    ws.cell(row=head_row, column=1, value="Tháng").font = Font(bold=True, size=9, color=INK2)
    n = len(names)
    for j, name in enumerate(names):
        c = ws.cell(row=head_row, column=2 + j, value=name)
        c.font = Font(bold=True, size=9, color=INK2)
    ws.cell(row=head_row, column=2 + n, value="Tổng").font = Font(bold=True, size=9, color=INK2)

    r1 = head_row + 1
    for m in range(1, 13):
        r = head_row + m
        ws.cell(row=r, column=1, value=f"Th{m}")
        for j, src_r in enumerate(src_rows):
            f = (f"=IF(AND({m}>=$B$1,{m}<=$C$1),"
                 f"'{SRC_SHEET}'!{col_m(m)}{src_r}/1000000,\"\")")
            cell = ws.cell(row=r, column=2 + j, value=f)
            cell.number_format = unit_fmt
        tc = ws.cell(row=r, column=2 + n,
                     value=f"=SUM({get_column_letter(2)}{r}:{get_column_letter(1 + n)}{r})")
        tc.number_format = unit_fmt
    r_tot = head_row + 13
    ws.cell(row=r_tot, column=1, value="Tổng kỳ lọc").font = Font(bold=True, size=9)
    for j in range(n + 1):
        col = get_column_letter(2 + j)
        cell = ws.cell(row=r_tot, column=2 + j, value=f"=SUM({col}{r1}:{col}{head_row + 12})")
        cell.number_format = unit_fmt
        cell.font = Font(bold=True, size=9)
    return r1, head_row + 12, r_tot


def build_db(ws, spdv_names, th_names, ton_names):
    ws.sheet_properties.tabColor = GRAY
    ws["A1"] = "Bộ lọc (Từ / Đến tháng) →"
    ws["A1"].font = Font(size=9, color=INK2)
    ws["B1"] = f"='{DASH_SHEET}'!C4"
    ws["C1"] = f"='{DASH_SHEET}'!E4"
    for a in ("B1", "C1"):
        ws[a].font = Font(bold=True)
        ws[a].fill = PatternFill("solid", fgColor=INPUT_FILL)
    ws["E1"] = "Sheet trung gian cho biểu đồ — có thể ẩn. Đơn vị: TRIỆU ₫."
    ws["E1"].font = Font(size=9, italic=True, color=INK2)

    ws.column_dimensions["A"].width = 12
    for c in range(2, 13):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Khối 1: CO NHẬP theo SPDV theo tháng (hàng 4..17)
    month_block(ws, "1) CO NHẬP THEO SPDV THEO THÁNG (triệu ₫)", 4,
                spdv_names, ROWS_NHAP, "#,##0")
    # Khối 2: CO XUẤT (hàng 20..33)
    month_block(ws, "2) CO XUẤT THEO SPDV THEO THÁNG (triệu ₫)", 20,
                spdv_names, ROWS_XUAT, "#,##0")
    # Khối 3: tỉ trọng CO NHẬP theo SPDV — kỳ lọc (hàng 36..45)
    ws.cell(row=35, column=1, value="3) TỈ TRỌNG CO NHẬP THEO SPDV — KỲ LỌC").font = \
        Font(bold=True, size=11, color=INK)
    ws.cell(row=36, column=1, value="SPDV").font = Font(bold=True, size=9, color=INK2)
    ws.cell(row=36, column=2, value="CO Nhập (triệu ₫)").font = Font(bold=True, size=9, color=INK2)
    for j, name in enumerate(spdv_names):
        ws.cell(row=37 + j, column=1, value=name)
        cell = ws.cell(row=37 + j, column=2, value=f"={get_column_letter(2 + j)}17")
        cell.number_format = "#,##0"
    # Khối 4: THIỆT HẠI theo nhóm theo tháng (hàng 48..61)
    month_block(ws, "4) THIỆT HẠI THEO SPDV THEO THÁNG (triệu ₫)", 48,
                th_names, ROWS_TH, "#,##0.0")
    # Khối 5: CO TỒN theo nhóm theo tháng (hàng 64..77)
    month_block(ws, "5) CO TỒN THEO SPDV THEO THÁNG (triệu ₫)", 64,
                ton_names, ROWS_TON, "#,##0.0")


# =====================================================================
# Lớp 3: trang 'Tổng quan vận hành'
# =====================================================================
def kpi_card(ws, row, col, label, formula, note=None, num_fmt='#,##0 "₫"'):
    c1, c2 = get_column_letter(col), get_column_letter(col + 1)
    ws.merge_cells(f"{c1}{row}:{c2}{row}")
    ws.merge_cells(f"{c1}{row + 1}:{c2}{row + 1}")
    lab = ws.cell(row=row, column=col, value=label)
    lab.font = Font(bold=True, size=9, color=INK2)
    lab.alignment = Alignment(horizontal="center", vertical="center")
    val = ws.cell(row=row + 1, column=col, value=formula)
    val.font = Font(bold=True, size=13, color=INK)
    val.number_format = num_fmt
    val.alignment = Alignment(horizontal="center", vertical="center")
    for rr in (row, row + 1):
        for cc in (col, col + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = PatternFill("solid", fgColor=CARD_FILL)
            cell.border = box
    if note:
        nt = ws.cell(row=row + 2, column=col, value=note)
        nt.font = Font(size=8, italic=True, color=INK2)
        nt.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{c1}{row + 2}:{c2}{row + 2}")


def sum_filtered(total_row):
    """SUMPRODUCT hàng tổng trong sheet nguồn theo bộ lọc C4/E4 (đơn vị ₫)."""
    rng = f"'{SRC_SHEET}'!$F${total_row}:$Q${total_row}"
    colx = f"COLUMN({rng})-5"
    return f"=SUMPRODUCT({rng},({colx}>=$C$4)*({colx}<=$E$4))"


def style_bar(chart, colors, title):
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 60
    chart.title = title
    chart.legend.position = "b"
    chart.y_axis.numFmt = "#,##0"
    chart.y_axis.title = "triệu ₫"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.width = 14.5
    chart.height = 8.2
    for s, color in zip(chart.series, colors):
        gp = GraphicalProperties(solidFill=color)
        gp.line = LineProperties(solidFill="FFFFFF", w=9525)  # khe 0.75pt giữa các tầng
        s.graphicalProperties = gp


def build_dash(ws, db_ws, spdv_names):
    ws.sheet_properties.tabColor = BLUE
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for c in range(2, 19):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["B1"] = "TỔNG QUAN VẬN HÀNH — PCU 2026"
    ws["B1"].font = Font(bold=True, size=16, color=INK)
    ws["B2"] = ("Nguồn: 'Data theo SPDV'  •  KPI: ₫  •  Biểu đồ & bảng: triệu ₫  •  "
                "Đổi Từ tháng/Đến tháng ở ô C4, E4 — toàn bộ KPI, biểu đồ, bảng cập nhật theo.")
    ws["B2"].font = Font(size=9, italic=True, color=INK2)

    # ---- Bộ lọc thời gian
    ws["B4"] = "Từ tháng"
    ws["D4"] = "Đến tháng"
    for a in ("B4", "D4"):
        ws[a].font = Font(bold=True, size=10)
        ws[a].alignment = Alignment(horizontal="right")
    ws["C4"] = 1
    ws["E4"] = 12
    for a in ("C4", "E4"):
        ws[a].font = Font(bold=True, size=12, color="1C5CAB")
        ws[a].fill = PatternFill("solid", fgColor=INPUT_FILL)
        ws[a].border = box
        ws[a].alignment = Alignment(horizontal="center")
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="12",
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Tháng 1–12", error="Nhập số tháng từ 1 đến 12",
                        showInputMessage=True, promptTitle="Bộ lọc tháng",
                        prompt="Nhập số 1–12")
    ws.add_data_validation(dv)
    dv.add(ws["C4"])
    dv.add(ws["E4"])
    ws["G4"] = "← Mặc định 1–12: cái nhìn tổng cả năm; thu hẹp để soi từng giai đoạn."
    ws["G4"].font = Font(size=9, italic=True, color=INK2)

    # ---- 5 thẻ KPI (hàng 6–8)
    kpi_card(ws, 6, 2, "TỔNG CO NHẬP (kỳ lọc)", sum_filtered(ROW_NHAP_TONG))
    kpi_card(ws, 6, 5, "TỔNG CO XUẤT (kỳ lọc)", sum_filtered(ROW_XUAT_TONG))
    kpi_card(ws, 6, 8, "CHÊNH LỆCH NHẬP − XUẤT", "=B7-E7")
    kpi_card(ws, 6, 11, "TỔNG THIỆT HẠI (kỳ lọc)", sum_filtered(ROW_TH_TONG),
             note='=IF(B7=0,"",TEXT(K7/B7,"0.00%")&" trên CO Nhập")')
    kpi_card(ws, 6, 14, "TỔNG CO TỒN (kỳ lọc)", sum_filtered(ROW_TON_TONG),
             note="cộng dồn các tháng trong kỳ")

    # ---- 5 biểu đồ
    # 1) CO Nhập stacked column
    c1 = BarChart()
    c1.add_data(Reference(db_ws, min_col=2, max_col=10, min_row=4, max_row=16),
                titles_from_data=True)
    c1.set_categories(Reference(db_ws, min_col=1, min_row=5, max_row=16))
    style_bar(c1, SPDV_COLORS, "1) Tổng CO NHẬP theo SPDV — theo tháng (triệu ₫)")
    ws.add_chart(c1, "B10")

    # 2) CO Xuất stacked column
    c2 = BarChart()
    c2.add_data(Reference(db_ws, min_col=2, max_col=10, min_row=20, max_row=32),
                titles_from_data=True)
    c2.set_categories(Reference(db_ws, min_col=1, min_row=21, max_row=32))
    style_bar(c2, SPDV_COLORS, "2) Tổng CO XUẤT theo SPDV — theo tháng (triệu ₫)")
    ws.add_chart(c2, "L10")

    # 3) Tỉ trọng CO Nhập — pie
    c3 = PieChart()
    c3.add_data(Reference(db_ws, min_col=2, min_row=36, max_row=45), titles_from_data=True)
    c3.set_categories(Reference(db_ws, min_col=1, min_row=37, max_row=45))
    c3.title = "3) Tỉ trọng CO NHẬP theo SPDV — kỳ lọc"
    c3.dataLabels = DataLabelList(showPercent=True, showVal=False, showCatName=False,
                                  showSerName=False, showLegendKey=False, showBubbleSize=False)
    c3.legend.position = "r"
    c3.width = 14.5
    c3.height = 8.2
    ser = c3.series[0]
    pts = []
    for i, color in enumerate(SPDV_COLORS):
        gp = GraphicalProperties(solidFill=color)
        gp.line = LineProperties(solidFill="FFFFFF", w=9525)
        pts.append(DataPoint(idx=i, spPr=gp))
    ser.data_points = pts
    ws.add_chart(c3, "B27")

    # 4) Thiệt hại stacked column
    c4 = BarChart()
    c4.add_data(Reference(db_ws, min_col=2, max_col=6, min_row=48, max_row=60),
                titles_from_data=True)
    c4.set_categories(Reference(db_ws, min_col=1, min_row=49, max_row=60))
    style_bar(c4, TH_COLORS, "4) Tổng THIỆT HẠI theo SPDV — theo tháng (triệu ₫)")
    c4.y_axis.numFmt = "#,##0.0"
    ws.add_chart(c4, "L27")

    # 5) CO Tồn stacked column
    c5 = BarChart()
    c5.add_data(Reference(db_ws, min_col=2, max_col=4, min_row=64, max_row=76),
                titles_from_data=True)
    c5.set_categories(Reference(db_ws, min_col=1, min_row=65, max_row=76))
    style_bar(c5, TON_COLORS, "5) Tổng CO TỒN theo SPDV — theo tháng (triệu ₫)")
    c5.y_axis.numFmt = "#,##0.0"
    ws.add_chart(c5, "B44")

    # ---- Bảng so sánh tháng (cạnh biểu đồ 5) — cũng là "table view" cho accessibility
    ws["L43"] = "SO SÁNH THÁNG (kỳ lọc, triệu ₫)"
    ws["L43"].font = Font(bold=True, size=11)
    heads = ["Tháng", "CO Nhập", "% MoM Nhập", "CO Xuất", "Thiệt hại", "CO Tồn"]
    for j, h in enumerate(heads):
        cell = ws.cell(row=44, column=12 + j, value=h)
        cell.font = Font(bold=True, size=9, color=INK2)
        cell.border = box
        cell.fill = PatternFill("solid", fgColor=CARD_FILL)
    for m in range(1, 13):
        r = 44 + m
        in_filter = f"AND({m}>=$C$4,{m}<=$E$4)"
        ws.cell(row=r, column=12, value=f'=IF({in_filter},"Th{m}","")')
        ws.cell(row=r, column=13, value=f"=IF({in_filter},'{DB_SHEET}'!K{4 + m},\"\")")
        if m == 1:
            ws.cell(row=r, column=14, value="")
        else:
            ws.cell(row=r, column=14,
                    value=f'=IF(OR(M{r - 1}="",M{r}="",M{r - 1}=0),"",M{r}/M{r - 1}-1)')
        ws.cell(row=r, column=15, value=f"=IF({in_filter},'{DB_SHEET}'!K{20 + m},\"\")")
        ws.cell(row=r, column=16, value=f"=IF({in_filter},'{DB_SHEET}'!G{48 + m},\"\")")
        ws.cell(row=r, column=17, value=f"=IF({in_filter},'{DB_SHEET}'!E{64 + m},\"\")")
        for j, fmt in ((13, "#,##0"), (14, "+0.0%;-0.0%"), (15, "#,##0"),
                       (16, "#,##0.0"), (17, "#,##0.0")):
            cell = ws.cell(row=r, column=j)
            cell.number_format = fmt
            cell.border = box
        ws.cell(row=r, column=12).border = box
    r = 57
    ws.cell(row=r, column=12, value="Tổng").font = Font(bold=True, size=10)
    for j, ref, fmt in ((13, f"'{DB_SHEET}'!K17", "#,##0"),
                        (15, f"'{DB_SHEET}'!K33", "#,##0"),
                        (16, f"'{DB_SHEET}'!G61", "#,##0.0"),
                        (17, f"'{DB_SHEET}'!E77", "#,##0.0")):
        cell = ws.cell(row=r, column=j, value=f"={ref}")
        cell.font = Font(bold=True, size=10)
        cell.number_format = fmt
        cell.border = box
    ws.cell(row=r, column=12).border = box
    ws.cell(row=r, column=14).border = box

    ws.freeze_panes = "A9"


def main(src_path, out_path):
    print("Đọc dữ liệu nguồn…")
    src_wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    src_ws = src_wb[SRC_SHEET]
    values = list(src_ws.iter_rows(min_row=1, max_row=50, max_col=18, values_only=True))
    src_wb.close()

    def name_at(r):
        return values[r - 1][3]  # cột D

    spdv_names = [name_at(r) for r in ROWS_NHAP]
    th_names = [name_at(r) for r in ROWS_TH]
    ton_names = [name_at(r) for r in ROWS_TON]
    print("SPDV:", spdv_names)
    print("Thiệt hại:", th_names)
    print("Tồn:", ton_names)

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = DASH_SHEET
    ws_db = wb.create_sheet(DB_SHEET)
    ws_src = wb.create_sheet(SRC_SHEET)

    copy_snapshot(values, ws_src)
    build_db(ws_db, spdv_names, th_names, ton_names)
    build_dash(ws_dash, ws_db, spdv_names)

    wb.save(out_path)
    print("Đã ghi:", out_path)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
